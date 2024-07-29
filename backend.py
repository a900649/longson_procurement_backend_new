from datetime import datetime,timezone,timedelta
import pandas as pd
import streamlit as st
import numpy as np
import warnings
import json
from datetime import datetime
from io import BytesIO
import openpyxl as op
from openpyxl.styles import Font,NamedStyle,Alignment
import warnings
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth
from azure.storage.blob import BlobServiceClient,ContainerClient
import os
import math
import variable as v
import my_function as my

warnings.filterwarnings("ignore")
st.set_page_config(page_title=v.page_title, page_icon=v.icon_path)

with open("user.yaml") as file:
    config = yaml.load(file, Loader=SafeLoader)

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days'],
    config['preauthorized'],
)

def create_project():
    st.title("Create Project")
    st.markdown('---')

    project_name_to_data, project_id_to_data = my.get_project_dict()
    title_name_to_id_dict, title_id_to_name_dict = my.get_title_name_to_id_dict()
    requirements_introduction_name_to_id_dict, requirements_introduction_id_to_name_dict = my.get_requirements_introduction_name_to_id_dict()
    common_columns_name_to_id_dict, common_columns_id_to_name_dict = my.get_common_columns_name_to_id_dict()
    individual_columns_name_to_id_dict, individual_columns_id_to_name_dict = my.get_individual_columns_name_to_id_dict()
    function_name_to_id_dict, function_id_to_name_dict = my.get_function_name_to_id_dict()
    verification_code_name_to_id_dict, verification_code_id_to_name_dict = my.get_verification_code_name_to_id_dict()
    internal_mail_receipients_name_to_id_dict, internal_mail_receipients_id_to_name_dict = my.get_internal_mail_receipients_name_to_id_dict()
    supplier_mail_setting_name_to_id_dict, supplier_mail_setting_id_to_name_dict = my.get_supplier_mail_setting_name_to_id_dict()
    check_name_to_id_dict, check_id_to_name_dict = my.get_check_name_to_id_dict()

    tab1, tab2, tab3, tab4 = st.tabs(["Show","Create","Edit Name","Edit Content"])

    radio_option = ["True", "False"]

    with tab1:
        if st.button('Refresh', key="Refresh Show Project"):
            st.rerun()
        select_project = st.selectbox("Select Project", list(project_name_to_data.keys()), key="Show Project Selectbox")

        if select_project != None:
            st.selectbox("Select Title", list(title_name_to_id_dict.keys()), key="Show Project Title Selectbox", disabled=True, index=list(title_name_to_id_dict.keys()).index(title_id_to_name_dict[project_name_to_data[select_project][2]]))
            st.selectbox("Select Requirements Introduction", list(requirements_introduction_name_to_id_dict.keys()), key="Show Project Requirements Introduction Selectbox", disabled=True, index=list(requirements_introduction_name_to_id_dict.keys()).index(requirements_introduction_id_to_name_dict[project_name_to_data[select_project][3]]))
            st.selectbox("Select Common Columns", list(common_columns_name_to_id_dict.keys()), key="Show Project Common Columns Selectbox", disabled=True, index=list(common_columns_name_to_id_dict.keys()).index(common_columns_id_to_name_dict[project_name_to_data[select_project][4]]))
            st.selectbox("Select Individual Columns", list(individual_columns_name_to_id_dict.keys()), key="Show Project Individual Columns Selectbox", disabled=True, index=list(individual_columns_name_to_id_dict.keys()).index(individual_columns_id_to_name_dict[project_name_to_data[select_project][5]]))
            st.selectbox("Select Function", list(function_name_to_id_dict.keys()), key="Show Project Function Selectbox", disabled=True,index=list(function_name_to_id_dict.keys()).index(function_id_to_name_dict[project_name_to_data[select_project][6]]))
            st.selectbox("Select Verification Code", list(verification_code_name_to_id_dict.keys()), key="Show Project Verification Code Selectbox", disabled=True, index=list(verification_code_name_to_id_dict.keys()).index(verification_code_id_to_name_dict[project_name_to_data[select_project][7]]))
            st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_name_to_id_dict.keys()), key="Show Project Internal Mail Receipients Selectbox", disabled=True, index=list(internal_mail_receipients_name_to_id_dict.keys()).index(internal_mail_receipients_id_to_name_dict[project_name_to_data[select_project][8]]))
            st.selectbox("Select Supplier Mail Setting", list(supplier_mail_setting_name_to_id_dict.keys()), key="Show Project Supplier Mail Setting Selectbox", disabled=True, index=list(supplier_mail_setting_name_to_id_dict.keys()).index(supplier_mail_setting_id_to_name_dict[project_name_to_data[select_project][9]]))
            st.selectbox("Select Check", list(check_name_to_id_dict.keys()), key="Show Project Check Selectbox", disabled=True, index=list(check_name_to_id_dict.keys()).index(check_id_to_name_dict[project_name_to_data[select_project][10]]))
            st.radio("Select Status", radio_option, key="Show Start Radio", disabled=True,index=radio_option.index(project_name_to_data[select_project][11]))

    with tab2:
        new_name = st.text_input("New Project Name", key="Add Project Name")

        select_title = st.selectbox("Select Title", list(title_name_to_id_dict.keys()), key="Add Project Title Selectbox")
        select_requirements_introduction = st.selectbox("Select Requirements Introduction", list(requirements_introduction_name_to_id_dict.keys()), key="Add Project Requirements Introduction Selectbox")
        select_common_columns = st.selectbox("Select Common Columns", list(common_columns_name_to_id_dict.keys()), key="Add Project Common Columns Selectbox")
        select_individual_columns = st.selectbox("Select Individual Columns", list(individual_columns_name_to_id_dict.keys()), key="Add Project Individual Columns Selectbox")
        select_function = st.selectbox("Select Function", list(function_name_to_id_dict.keys()), key="Add Project Function Selectbox")
        select_verification_code = st.selectbox("Select Verification Code", list(verification_code_name_to_id_dict.keys()), key="Add Project Verification Code Selectbox")
        select_internal_mail_receipients = st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_name_to_id_dict.keys()), key="Add Project Internal Mail Receipients Selectbox")
        select_supplier_mail_setting = st.selectbox("Select Supplier Mail Setting", list(supplier_mail_setting_name_to_id_dict.keys()), key="Add Project Supplier Mail Setting Selectbox")
        select_check = st.selectbox("Select Check", list(check_name_to_id_dict.keys()), key="Add Project Check Selectbox")
        select_start = st.radio("Select Status", radio_option, key="Add Start Radio")

        if st.button('Confirm', key="Add Project Name Button"):
            if new_name in list(project_name_to_data.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_project(new_name,title_name_to_id_dict[select_title],requirements_introduction_name_to_id_dict[select_requirements_introduction],
                                  common_columns_name_to_id_dict[select_common_columns],individual_columns_name_to_id_dict[select_individual_columns],
                                  function_name_to_id_dict[select_function],verification_code_name_to_id_dict[select_verification_code],
                                  internal_mail_receipients_name_to_id_dict[select_internal_mail_receipients],
                                  supplier_mail_setting_name_to_id_dict[select_supplier_mail_setting],check_name_to_id_dict[select_check],select_start)
                project_name_to_data, project_id_to_data = my.get_project_dict()
                st.success("Added Successfully")

    with tab3:

        if st.button('Refresh', key="Refresh Edit Project Name"):
            st.rerun()
        select_project = st.selectbox("Select Project", list(project_name_to_data.keys()), key="Edit Project Name Selectbox")

        edit_name = st.text_input("New Project Name", key="Edit Project Name")

        if select_project != None:
            if st.button('Confirm', key="Edit Title Name"):
                if edit_name in list(project_name_to_data.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_project_name(select_project,edit_name)
                    project_name_to_data, project_id_to_data = my.get_project_dict()
                    st.success("Change Successful")

    with tab4:

        if st.button('Refresh', key="Refresh Edit Project Content"):
            st.rerun()
        select_project = st.selectbox("Select Project", list(project_name_to_data.keys()), key="Edit Project Content Selectbox")

        if select_project != None:
            select_title = st.selectbox("Select Title", list(title_name_to_id_dict.keys()), key="Edit Project Title Selectbox",index=list(title_name_to_id_dict.keys()).index(title_id_to_name_dict[project_name_to_data[select_project][2]]))
            select_requirements_introduction = st.selectbox("Select Requirements Introduction", list(requirements_introduction_name_to_id_dict.keys()), key="Edit Project Requirements Introduction Selectbox", index=list(requirements_introduction_name_to_id_dict.keys()).index(requirements_introduction_id_to_name_dict[project_name_to_data[select_project][3]]))
            select_common_columns = st.selectbox("Select Common Columns", list(common_columns_name_to_id_dict.keys()), key="Edit Project Common Columns Selectbox", index=list(common_columns_name_to_id_dict.keys()).index(common_columns_id_to_name_dict[project_name_to_data[select_project][4]]))
            select_individual_columns = st.selectbox("Select Individual Columns", list(individual_columns_name_to_id_dict.keys()), key="Edit Project Individual Columns Selectbox", index=list(individual_columns_name_to_id_dict.keys()).index(individual_columns_id_to_name_dict[project_name_to_data[select_project][5]]))
            select_function = st.selectbox("Select Function", list(function_name_to_id_dict.keys()), key="Edit Project Function Selectbox", index=list(function_name_to_id_dict.keys()).index(function_id_to_name_dict[project_name_to_data[select_project][6]]))
            select_verification_code = st.selectbox("Select Verification Code", list(verification_code_name_to_id_dict.keys()), key="Edit Project Verification Code Selectbox", index=list(verification_code_name_to_id_dict.keys()).index(verification_code_id_to_name_dict[project_name_to_data[select_project][7]]))
            select_internal_mail_receipients = st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_name_to_id_dict.keys()), key="Edit Project Internal Mail Receipients Selectbox", index=list(internal_mail_receipients_name_to_id_dict.keys()).index(internal_mail_receipients_id_to_name_dict[project_name_to_data[select_project][8]]))
            select_supplier_mail_setting = st.selectbox("Select Supplier Mail Setting", list(supplier_mail_setting_name_to_id_dict.keys()), key="Edit Project Supplier Mail Setting Selectbox", index=list(supplier_mail_setting_name_to_id_dict.keys()).index(supplier_mail_setting_id_to_name_dict[project_name_to_data[select_project][9]]))
            select_check = st.selectbox("Select Check", list(check_name_to_id_dict.keys()), key="Edit Project Check Selectbox", index=list(check_name_to_id_dict.keys()).index(check_id_to_name_dict[project_name_to_data[select_project][10]]))
            select_start = st.radio("Select Status", radio_option, key="Edit Start Radio", index=radio_option.index(project_name_to_data[select_project][11]))

            if st.button('Confirm', key="Edit Project Content"):

                my.edit_project_row(select_project,title_name_to_id_dict[select_title],requirements_introduction_name_to_id_dict[select_requirements_introduction],
                                    common_columns_name_to_id_dict[select_common_columns],individual_columns_name_to_id_dict[select_individual_columns],
                                    function_name_to_id_dict[select_function],verification_code_name_to_id_dict[select_verification_code],
                                    internal_mail_receipients_name_to_id_dict[select_internal_mail_receipients],supplier_mail_setting_name_to_id_dict[select_supplier_mail_setting],
                                    check_name_to_id_dict[select_check],select_start)
                project_name_to_data, project_id_to_data = my.get_project_dict()
                st.success("Change Successful")


def setting_title():
    st.title("Setting Title")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    title_dict = my.get_title_dict()

    with tab1:
        data_list = []
        for key in title_dict.keys():
            data = [key] + title_dict[key]
            data_list.append(data)

        if st.button('Refresh', key="Refresh Show Title"):
            st.rerun()

        title_df = pd.DataFrame(data_list, columns=["Title", "Title Row 1", "Title Row 2", "Title Row 3", "Title Row 4", "Title Row 5"],
                                index=[i + 1 for i in range(0, len(data_list))])


        st.dataframe(title_df)

    with tab2:
        new_title_name = st.text_input("New Title Name")
        title_row_1 = st.text_input("Title Row 1")
        title_row_2 = st.text_input("Title Row 2")
        title_row_3 = st.text_input("Title Row 3")
        title_row_4 = st.text_input("Title Row 4")
        title_row_5 = st.text_input("Title Row 5")

        if st.button('Confirm', key="Add Title"):
            if new_title_name in list(title_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_title(new_title_name,title_row_1,title_row_2,title_row_3,title_row_4,title_row_5)
                title_dict = my.get_title_dict()
                st.success("Added Successfully")

    with tab3:
        select_title = st.selectbox("Select Title", list(title_dict.keys()),key="Edit Title Name Selectbox")

        edit_title = st.text_input("New Title")

        if st.button('Refresh', key="Refresh Edit Title Name"):
            st.rerun()

        if select_title != None:
            if st.button('Confirm', key="Edit Title"):
                if edit_title in list(title_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_title_name(select_title,edit_title)
                    title_dict = my.get_title_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Title Content"):
            st.rerun()

        select_title = st.selectbox("Select Title", list(title_dict.keys()),key="Edit Title Content Selectbox")
        if select_title != None:
            title_row_1 = st.text_input("Title Row 1",title_dict[select_title][0],key="Edit Title Row 1")
            title_row_2 = st.text_input("Title Row 2",title_dict[select_title][1],key="Edit Title Row 2")
            title_row_3 = st.text_input("Title Row 3",title_dict[select_title][2],key="Edit Title Row 3")
            title_row_4 = st.text_input("Title Row 4",title_dict[select_title][3],key="Edit Title Row 4")
            title_row_5 = st.text_input("Title Row 5",title_dict[select_title][4],key="Edit Title Row 5")

            if st.button('Confirm', key="Edit Title Content"):
                my.edit_title_row(select_title, title_row_1, title_row_2, title_row_3, title_row_4, title_row_5)
                title_dict = my.get_title_dict()
                st.success("Change Successful")

def setting_requirements_introduction():
    st.title("Requirements Introduction")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    requirements_introduction_dict = my.get_requirements_introduction_dict()

    with tab1:
        if st.button('Refresh', key="Refresh Show Requirements Introduction"):
            st.rerun()
        select_requirements_introduction = st.selectbox("Select Requirements Introduction", list(requirements_introduction_dict.keys()), key="Show Requirements Introduction Selectbox")
        if select_requirements_introduction != None:
            st.text_input("Product 1",requirements_introduction_dict[select_requirements_introduction][0],disabled=True,key="Show Requirements Introduction Product 1")
            st.text_area("Introduction 1",requirements_introduction_dict[select_requirements_introduction][1],disabled=True,key="Show Requirements Introduction Introduction 1")
            st.text_input("Product 2",requirements_introduction_dict[select_requirements_introduction][2],disabled=True,key="Show Requirements Introduction Product 2")
            st.text_area("Introduction 2",requirements_introduction_dict[select_requirements_introduction][3],disabled=True,key="Show Requirements Introduction Introduction 2")
            st.text_input("Product 3",requirements_introduction_dict[select_requirements_introduction][4],disabled=True,key="Show Requirements Introduction Product 3")
            st.text_area("Introduction 3",requirements_introduction_dict[select_requirements_introduction][5],disabled=True,key="Show Requirements Introduction Introduction 3")
            st.text_input("Product 4",requirements_introduction_dict[select_requirements_introduction][6],disabled=True,key="Show Requirements Introduction Product 4")
            st.text_area("Introduction 4",requirements_introduction_dict[select_requirements_introduction][7],disabled=True,key="Show Requirements Introduction Introduction 4")
            st.text_input("Product 5",requirements_introduction_dict[select_requirements_introduction][8],disabled=True,key="Show Requirements Introduction Product 5")
            st.text_area("Introduction 5",requirements_introduction_dict[select_requirements_introduction][9],disabled=True,key="Show Requirements Introduction Introduction 5")

    with tab2:
        new_requirements_introduction_name = st.text_input("Requirements Introduction Name")
        product_1 = st.text_input("Product 1")
        introduction_1 = st.text_area("Introduction 1")
        product_2 = st.text_input("Product 2")
        introduction_2 = st.text_area("Introduction 2")
        product_3 = st.text_input("Product 3")
        introduction_3 = st.text_area("Introduction 3")
        product_4 = st.text_input("Product 4")
        introduction_4 = st.text_area("Introduction 4")
        product_5 = st.text_input("Product 5")
        introduction_5 = st.text_area("Introduction 5")

        if st.button('Confirm', key="Add Requirements Introduction"):
            if new_requirements_introduction_name in list(requirements_introduction_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_requirements_introduction(new_requirements_introduction_name,product_1,introduction_1,product_2,introduction_2,
                                                    product_3,introduction_3,product_4,introduction_4,product_5,introduction_5)
                requirements_introduction_dict = my.get_requirements_introduction_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Requirements Introduction Name"):
            st.rerun()
        select_requirements_introduction_name = st.selectbox("Select Requirements Introduction", list(requirements_introduction_dict.keys()), key="Edit Requirements Introduction Name Selectbox")

        edit_requirements_introduction_name = st.text_input("New Requirements Introduction Name")

        if select_requirements_introduction_name != None:
            if st.button('Confirm', key="Edit Requirements Introduction Name"):
                if edit_requirements_introduction_name in list(requirements_introduction_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_requirements_introduction_name(select_requirements_introduction_name,edit_requirements_introduction_name)
                    requirements_introduction_dict = my.get_requirements_introduction_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Requirements Introduction Content"):
            st.rerun()
        select_requirements_introduction = st.selectbox("Select Requirements Introduction", list(requirements_introduction_dict.keys()), key="Edit Requirements Introduction Content Selectbox")
        if select_requirements_introduction != None:
            product_1 = st.text_input("Product 1",requirements_introduction_dict[select_requirements_introduction][0],key="Edit Requirements Introduction Content Product 1")
            introduction_1 = st.text_area("Introduction 1",requirements_introduction_dict[select_requirements_introduction][1],key="Edit Requirements Introduction Content Introduction 1")
            product_2 = st.text_input("Product 2",requirements_introduction_dict[select_requirements_introduction][2],key="Edit Requirements Introduction Content Product 2")
            introduction_2 = st.text_area("Introduction 2",requirements_introduction_dict[select_requirements_introduction][3],key="Edit Requirements Introduction Content Introduction 2")
            product_3 = st.text_input("Product 3",requirements_introduction_dict[select_requirements_introduction][4],key="Edit Requirements Introduction Content Product 3")
            introduction_3 = st.text_area("Introduction 3",requirements_introduction_dict[select_requirements_introduction][5],key="Edit Requirements Introduction Content Introduction 3")
            product_4 = st.text_input("Product 4",requirements_introduction_dict[select_requirements_introduction][6],key="Edit Requirements Introduction Content Product 4")
            introduction_4 = st.text_area("Introduction 4",requirements_introduction_dict[select_requirements_introduction][7],key="Edit Requirements Introduction Content Introduction 4")
            product_5 = st.text_input("Product 5",requirements_introduction_dict[select_requirements_introduction][8],key="Edit Requirements Introduction Content Product 5")
            introduction_5 = st.text_area("Introduction 5",requirements_introduction_dict[select_requirements_introduction][9],key="Edit Requirements Introduction Content Introduction 5")

            if st.button('Confirm', key="Edit Requirements Introduction Content"):
                my.edit_requirements_introduction_row(select_requirements_introduction, product_1, introduction_1, product_2,
                                                      introduction_2, product_3, introduction_3, product_4, introduction_4, product_5, introduction_5)
                requirements_introduction_dict = my.get_requirements_introduction_dict()
                st.success("Change Successful")

def setting_common_columns():
    st.title("Common Columns")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    common_columns_dict = my.get_common_columns_dict()

    default_option_list = ["Text", "Integer", "Decimal", "Date", "Radio"]

    with tab1:
        if st.button('Refresh', key="Refresh Show Common Columns"):
            st.rerun()

        select_common_columns_name = st.selectbox("Select Common Columns", list(common_columns_dict.keys()), key="Show Common Columns Selectbox")

        if common_columns_dict != {} and select_common_columns_name != None:
            data_str = common_columns_dict[select_common_columns_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_common_columns_name != None:
                for i in range(0, 20):
                    a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
                    with a1:
                        st.text_input("Column Name",data_df.loc[i,"Column Name"], disabled=True, key="Show Common Column Name {}".format(i))
                    with a2:
                        st.selectbox("Data Type", default_option_list,index=default_option_list.index(data_df.loc[i,"Data Type"]), disabled=True, key="Show Common Columns Data Type {}".format(i))
                    with a3:
                        st.text_area("Option",data_df.loc[i,"Option"], disabled=True, key="Show Common Columns Option {}".format(i))
                    with a4:
                        st.text_input("Instructions",data_df.loc[i,"Instructions"], disabled=True, key="Show Common Columns Instructions {}".format(i))

    with tab2:
        column_name_list = []
        data_type_list = []
        option_list = []
        instructions_list = []

        new_common_column_name = st.text_input("Name", key="Create Common Column Name")

        for i in range(0,20):
            a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
            with a1:
                column_name_list.append(st.text_input("Column Name",key="Create Common Column Name {}".format(i)))
            with a2:
                data_type_list.append(st.selectbox("Data Type", default_option_list, key="Create Common Columns Data Type {}".format(i)))
            with a3:
                option_list.append(st.text_area("Option", key="Create Common Columns Option {}".format(i)))
            with a4:
                instructions_list.append(st.text_input("Instructions", key="Create Common Columns Instructions {}".format(i)))


        if st.button('Confirm', key="Add Common Columns"):
            if new_common_column_name in list(common_columns_dict.keys()):
                st.warning("Duplicate Name")
            else:
                # Option部分清除不需要的文字
                option_list = [i.strip("\n") for i in option_list]
                option_list = [i.strip(" ") for i in option_list]
                my.create_common_columns(new_common_column_name,column_name_list,data_type_list,option_list,instructions_list)
                common_columns_dict = my.get_common_columns_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Common Columns Name"):
            st.rerun()
        select_common_columns_name = st.selectbox("Select Common Columns", list(common_columns_dict.keys()), key="Edit Common Columns Name Selectbox")

        edit_common_columns_name = st.text_input("New Common Columns Name")

        if select_common_columns_name == None:
            if st.button('Confirm', key="Edit Common Columns Name"):
                if edit_common_columns_name in list(common_columns_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_common_columns_name(select_common_columns_name,edit_common_columns_name)
                    common_columns_dict = my.get_common_columns_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Common Columns Content"):
            st.rerun()
        select_common_columns_name = st.selectbox("Select Common Columns", list(common_columns_dict.keys()), key="Edit Common Columns Content Selectbox")
        if common_columns_dict != {} and select_common_columns_name != None:
            data_str = common_columns_dict[select_common_columns_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_common_columns_name != None:
                for i in range(0, 20):
                    a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
                    with a1:
                        data_df.loc[i, "Column Name"] = st.text_input("Column Name", data_df.loc[i, "Column Name"], key="Edit Common Columns Content Column Name {}".format(i))
                    with a2:
                        data_df.loc[i, "Data Type"] = st.selectbox("Data Type", default_option_list, index=default_option_list.index(data_df.loc[i, "Data Type"]), key="Edit Common Columns Content Data Type {}".format(i))
                    with a3:
                        data_df.loc[i, "Option"] = st.text_area("Option", data_df.loc[i, "Option"], key="Edit Common Columns Content Option {}".format(i))
                    with a4:
                        data_df.loc[i, "Instructions"] = st.text_input("Instructions", data_df.loc[i, "Instructions"], key="Edit Common Columns Content Instructions {}".format(i))

                if st.button('Confirm', key="Edit Common Columns Content"):
                    # Option部分清除不需要的文字
                    data_df["Option"] = [i.strip("\n") for i in data_df["Option"]]
                    data_df["Option"] = [i.strip("") for i in data_df["Option"]]
                    common_columns_data = data_df.to_json(orient="split")
                    my.edit_common_columns_row(select_common_columns_name, common_columns_data)
                    common_columns_dict = my.get_common_columns_dict()
                    st.success("Change Successful")

def setting_individual_columns():
    st.title("Individual Columns")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    individual_columns_dict = my.get_individual_columns_dict()

    default_option_list = ["Text", "Integer", "Decimal", "Date", "Radio"]

    with tab1:
        if st.button('Refresh', key="Refresh Show Individual Columns"):
            st.rerun()

        select_individual_columns_name = st.selectbox("Select Individual Columns", list(individual_columns_dict.keys()), key="Show Individual Columns Selectbox")

        if individual_columns_dict != {} and select_individual_columns_name != None:
            data_str = individual_columns_dict[select_individual_columns_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_individual_columns_name != None:
                for i in range(0, 50):
                    a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
                    with a1:
                        st.text_input("Column Name",data_df.loc[i,"Column Name"], disabled=True, key="Show Individual Column Name {}".format(i))
                    with a2:
                        st.selectbox("Data Type", default_option_list,index=default_option_list.index(data_df.loc[i,"Data Type"]), disabled=True, key="Show Individual Columns Data Type {}".format(i))
                    with a3:
                        st.text_area("Option",data_df.loc[i,"Option"], disabled=True, key="Show Individual Columns Option {}".format(i))
                    with a4:
                        st.text_input("Instructions",data_df.loc[i,"Instructions"], disabled=True, key="Show Individual Columns Instructions {}".format(i))

    with tab2:
        column_name_list = []
        data_type_list = []
        option_list = []
        instructions_list = []

        new_individual_column_name = st.text_input("Name", key="Create Individual Column Name")

        for i in range(0,50):
            a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
            with a1:
                column_name_list.append(st.text_input("Column Name",key="Create Individual Columns Name {}".format(i)))
            with a2:
                data_type_list.append(st.selectbox("Data Type", default_option_list, key="Create Individual Columns Data Type {}".format(i)))
            with a3:
                option_list.append(st.text_area("Option", key="Create Individual Columns Option {}".format(i)))
            with a4:
                instructions_list.append(st.text_input("Instructions", key="Create Individual Columns Instructions {}".format(i)))


        if st.button('Confirm', key="Add Individual Columns"):
            if new_individual_column_name in list(individual_columns_dict.keys()):
                st.warning("Duplicate Name")
            else:
                # Option部分清除不需要的文字
                option_list = [i.strip("\n") for i in option_list]
                option_list = [i.strip(" ") for i in option_list]
                my.create_individual_columns(new_individual_column_name,column_name_list,data_type_list,option_list,instructions_list)
                individual_columns_dict = my.get_individual_columns_dict()
                st.success("Added Successfully")


    with tab3:
        if st.button('Refresh', key="Refresh Edit Individual Columns Name"):
            st.rerun()
        select_individual_columns_name = st.selectbox("Select Individual Columns", list(individual_columns_dict.keys()), key="Edit Individual Columns Name Selectbox")

        edit_individual_columns_name = st.text_input("New Individual Columns Name")

        if select_individual_columns_name != None:
            if st.button('Confirm', key="Edit Individual Columns Name"):
                if edit_individual_columns_name in list(individual_columns_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_individual_columns_name(select_individual_columns_name,edit_individual_columns_name)
                    individual_columns_dict = my.get_individual_columns_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Individual Columns Content"):
            st.rerun()
        select_individual_columns_name = st.selectbox("Select Individual Columns", list(individual_columns_dict.keys()), key="Edit Individual Columns Content Selectbox")
        if individual_columns_dict != {} and select_individual_columns_name != None:
            data_str = individual_columns_dict[select_individual_columns_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_individual_columns_name != None:
                for i in range(0, 50):
                    a1, a2, a3, a4 = st.columns([1, 1, 1, 1])
                    with a1:
                        data_df.loc[i, "Column Name"] = st.text_input("Column Name", data_df.loc[i, "Column Name"], key="Edit Individual Columns Content Column Name {}".format(i))
                    with a2:
                        data_df.loc[i, "Data Type"] = st.selectbox("Data Type", default_option_list, index=default_option_list.index(data_df.loc[i, "Data Type"]), key="Edit Individual Columns Content Data Type {}".format(i))
                    with a3:
                        data_df.loc[i, "Option"] = st.text_area("Option", data_df.loc[i, "Option"], key="Edit Individual Columns Content Option {}".format(i))
                    with a4:
                        data_df.loc[i, "Instructions"] = st.text_input("Instructions", data_df.loc[i, "Instructions"], key="Edit Individual Columns Content Instructions {}".format(i))

                if st.button('Confirm', key="Edit Individual Columns Content"):
                    # Option部分清除不需要的文字
                    data_df["Option"] = [i.strip("\n") for i in data_df["Option"]]
                    data_df["Option"] = [i.strip("") for i in data_df["Option"]]
                    individual_columns_data = data_df.to_json(orient="split")
                    my.edit_individual_columns_row(select_individual_columns_name, individual_columns_data)
                    individual_columns_dict = my.get_individual_columns_dict()
                    st.success("Change Successful")

def setting_function():
    st.title("Function Columns")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    function_dict = my.get_function_dict()

    radio_option = ["True","False"]

    with tab1:
        if st.button('Refresh', key="Refresh Show Function"):
            st.rerun()
        select_function = st.selectbox("Select Function", list(function_dict.keys()), key="Show Function Selectbox")
        if select_function != None:
            data_list = function_dict[select_function]
            st.radio("Attachment", key="Show Function Attachment",disabled=True, options=radio_option, index=radio_option.index(data_list[0]))
            st.radio("Verification Code", key="Show Function Verification Code",disabled=True, options=radio_option, index=radio_option.index(data_list[1]))
            st.radio("Send Internal Mail", key="Show Function Send Internal Mail",disabled=True, options=radio_option, index=radio_option.index(data_list[2]))
            st.radio("Send Supplier Mail", key="Show Function Send Supplier Mail",disabled=True, options=radio_option, index=radio_option.index(data_list[3]))

    with tab2:
        new_function_name = st.text_input("Function Name")
        attachment = st.radio("Attachment", key="Add Function Attachment", options=radio_option, index=1)
        verification_code = st.radio("Verification Code", key="Add Function Verification Code", options=radio_option, index=1)
        send_internal_mail = st.radio("Send Internal Mail", key="Add Function Send Internal Mail", options=radio_option, index=1)
        send_supplier_mail = st.radio("Send Supplier Mail", key="Add Function Send Supplier Mail", options=radio_option, index=1)

        if st.button('Confirm', key="Add Function"):
            if new_function_name in list(function_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_function(new_function_name,attachment,verification_code,send_internal_mail,send_supplier_mail)
                function_dict = my.get_function_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Function Name"):
            st.rerun()
        select_function = st.selectbox("Select Function", list(function_dict.keys()), key="Edit Function Name Selectbox")

        edit_function_name = st.text_input("Edit Function Name")

        if select_function != None:
            if st.button('Confirm', key="Edit Requirements Introduction Name"):
                if edit_function_name in list(function_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_function_name(select_function,edit_function_name)
                    function_dict = my.get_function_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Function Content"):
            st.rerun()
        select_function = st.selectbox("Select Function", list(function_dict.keys()), key="Edit Function Content Selectbox")
        if select_function != None:
            data_list = function_dict[select_function]
            attachment = st.radio("Attachment", key="Edit Function Attachment", options=radio_option, index=radio_option.index(data_list[0]))
            verification_code = st.radio("Verification Code", key="Edit Function Verification Code", options=radio_option, index=radio_option.index(data_list[1]))
            send_internal_mail = st.radio("Send Internal Mail", key="Edit Function Send Internal Mail", options=radio_option, index=radio_option.index(data_list[2]))
            send_supplier_mail = st.radio("Send Supplier Mail", key="Edit Function Send Supplier Mail", options=radio_option, index=radio_option.index(data_list[3]))

            if st.button('Confirm', key="Edit Requirements Introduction Content"):
                my.edit_function_row(select_function,attachment,verification_code,send_internal_mail,send_supplier_mail)
                function_dict = my.get_function_dict()
                st.success("Change Successful")

def setting_verification_code():
    st.title("Verification Code")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    verification_code_dict = my.get_verification_code_dict()

    with tab1:
        if st.button('Refresh', key="Refresh Show Verification Code"):
            st.rerun()

        select_verification_code_name = st.selectbox("Select Verification Code", list(verification_code_dict.keys()), key="Show Verification Code Selectbox")

        if verification_code_dict != {} and select_verification_code_name != None:
            data_str = verification_code_dict[select_verification_code_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_verification_code_name != None:
                for i in range(0, 100):
                    a1, a2 = st.columns([1, 1])
                    with a1:
                        st.text_input("Name",data_df.loc[i,"Name"], disabled=True, key="Show Verification Code Name {}".format(i))
                    with a2:
                        st.text_input("Code",data_df.loc[i,"Code"], disabled=True, key="Show Verification Code Code {}".format(i))

    with tab2:
        name_list = []
        code_list = []

        new_verification_code_name = st.text_input("Verification Code Name", key="Create Verification Code Name")

        for i in range(0,100):
            a1, a2 = st.columns([1, 1])
            with a1:
                name_list.append(st.text_input("Name",key="Create Verification Code Name {}".format(i)))
            with a2:
                code_list.append(st.text_input("Code",key="Create Verification Code Code {}".format(i)))

        if st.button('Confirm', key="Add Common Columns"):
            if new_verification_code_name in list(verification_code_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_verification_code(new_verification_code_name,name_list,code_list)
                verification_code_dict = my.get_verification_code_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Verification Code Name"):
            st.rerun()
        select_verification_code_name = st.selectbox("Select Verification Code", list(verification_code_dict.keys()), key="Edit Verification Code Name Selectbox")

        edit_verification_code_name = st.text_input("Edit Verification Code Name")

        if select_verification_code_name != None:
            if st.button('Confirm', key="Edit Requirements Introduction Name"):
                if edit_verification_code_name in list(verification_code_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_verification_code_name(select_verification_code_name,edit_verification_code_name)
                    verification_code_dict = my.get_verification_code_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Verification Code Content"):
            st.rerun()
        select_verification_code_name = st.selectbox("Select Verification Code", list(verification_code_dict.keys()), key="Edit Verification Code Content Selectbox")
        if verification_code_dict != {} and select_verification_code_name != None:
            data_str = verification_code_dict[select_verification_code_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_verification_code_name != None:
                for i in range(0, 100):
                    a1, a2 = st.columns([1, 1])
                    with a1:
                        st.text_input("Name", data_df.loc[i, "Name"], key="Edit Verification Code Content Name {}".format(i))
                    with a2:
                        st.text_input("Code", data_df.loc[i, "Code"], key="Edit Verification Code Content Code {}".format(i))

                if st.button('Confirm', key="Edit Verification Code Content"):
                    verification_code_data = data_df.to_json(orient="split")
                    my.edit_verification_code_row(select_verification_code_name, verification_code_data)
                    verification_code_dict = my.get_verification_code_dict()
                    st.success("Change Successful")

def setting_internal_mail_receipients():
    st.title("Internal Mail Receipients")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    internal_mail_receipients_dict = my.get_internal_mail_receipients_dict()

    with tab1:
        if st.button('Refresh', key="Refresh Show Internal Mail Receipients"):
            st.rerun()
        select_internal_mail_receipients_name = st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_dict.keys()), key="Show Internal Mail Receipients Selectbox")

        if internal_mail_receipients_dict != {} and select_internal_mail_receipients_name != None:
            data_str = internal_mail_receipients_dict[select_internal_mail_receipients_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_internal_mail_receipients_name != None:
                for i in range(0, 100):
                    a1, a2 = st.columns([1, 1])
                    with a1:
                        st.text_input("Mail",data_df.loc[i,"Mail"], disabled=True, key="Show Internal Mail Receipients Mail {}".format(i))
                    with a2:
                        st.text_input("Subject",data_df.loc[i,"Subject"], disabled=True, key="Show Internal Mail Receipients Subject {}".format(i))

    with tab2:
        mail_list = []
        subject_list = []

        new_internal_mail_receipients_name = st.text_input("Name", key="Create Internal Mail Receipients Name")

        for i in range(0,100):
            a1, a2 = st.columns([1, 1])
            with a1:
                mail_list.append(st.text_input("Name",key="Create Internal Mail Receipients Mail {}".format(i)))
            with a2:
                subject_list.append(st.text_input("Code",key="Create Internal Mail Receipients Subject {}".format(i)))

        if st.button('Confirm', key="Add Common Columns"):
            if new_internal_mail_receipients_name in list(internal_mail_receipients_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_internal_mail_receipients(new_internal_mail_receipients_name,mail_list,subject_list)
                internal_mail_receipients_dict = my.get_internal_mail_receipients_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Internal Mail Receipients Name"):
            st.rerun()
        select_internal_mail_receipients_name = st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_dict.keys()), key="Edit Internal Mail Receipients Name Selectbox")

        edit_internal_mail_receipients_name = st.text_input("Edit Internal Mail Receipients Name")

        if select_internal_mail_receipients_name != None:
            if st.button('Confirm', key="Edit Requirements Introduction Name"):
                if edit_internal_mail_receipients_name in list(internal_mail_receipients_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_internal_mail_receipients_name(select_internal_mail_receipients_name,edit_internal_mail_receipients_name)
                    internal_mail_receipients_dict = my.get_internal_mail_receipients_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Internal Mail Receipients Content"):
            st.rerun()
        select_internal_mail_receipients_name = st.selectbox("Select Internal Mail Receipients", list(internal_mail_receipients_dict.keys()), key="Edit Internal Mail Receipients Content Selectbox")
        if internal_mail_receipients_dict != {} and select_internal_mail_receipients_name != None:
            data_str = internal_mail_receipients_dict[select_internal_mail_receipients_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_internal_mail_receipients_name != None:
                for i in range(0, 100):
                    a1, a2 = st.columns([1, 1])
                    with a1:
                        st.text_input("Mail", data_df.loc[i, "Mail"], key="Edit Internal Mail Receipients Mail {}".format(i))
                    with a2:
                        st.text_input("Subject", data_df.loc[i, "Subject"], key="Edit Internal Mail Receipients Subject {}".format(i))

                if st.button('Confirm', key="Edit Internal Mail Receipients Content"):
                    internal_mail_receipients_data = data_df.to_json(orient="split")
                    my.edit_internal_mail_receipients_row(select_internal_mail_receipients_name, internal_mail_receipients_data)
                    internal_mail_receipients_dict = my.get_internal_mail_receipients_dict()
                    st.success("Change Successful")

def setting_supplier_mail():
    st.title("Supplier Mail Setting")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    supplier_mail_setting_dict = my.get_supplier_mail_setting_dict()

    with tab1:
        if st.button('Refresh', key="Refresh Show Supplier Mail Setting"):
            st.rerun()
        select_supplier_mail_setting_name = st.selectbox("Select Supplier Mail Setting", list(supplier_mail_setting_dict.keys()), key="Show Supplier Mail Setting Selectbox")

        if supplier_mail_setting_dict != {} and select_supplier_mail_setting_name != None:
            data_list = supplier_mail_setting_dict[select_supplier_mail_setting_name]

            if select_supplier_mail_setting_name != None:
                st.text_input("Subject", data_list[0], disabled=True, key="Show Supplier Mail Setting Subject")
                st.text_area("Content", data_list[1], disabled=True, key="Show Supplier Mail Setting Content")

    with tab2:
        new_supplier_mail_setting_name = st.text_input("Setting Supplier Mail Name")
        subject = st.text_input("Subject", key="Add Supplier Mail Setting Subject")
        content = st.text_area("Content", key="Add Supplier Mail Setting Verification Content")

        if st.button('Confirm', key="Add Function"):
            if new_supplier_mail_setting_name in list(supplier_mail_setting_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_supplier_mail_setting(new_supplier_mail_setting_name,subject,content)
                supplier_mail_setting_dict = my.get_supplier_mail_setting_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Supplier Mail Setting Name"):
            st.rerun()
        select_supplier_mail_setting = st.selectbox("Select Supplier Mail Setting", list(supplier_mail_setting_dict.keys()), key="Edit Supplier Mail Setting Name Selectbox")

        edit_supplier_mail_setting_name = st.text_input("Edit Supplier Mail Setting Name")

        if select_supplier_mail_setting != None:
            if st.button('Confirm', key="Edit Supplier Mail Setting Name"):
                if edit_supplier_mail_setting_name in list(supplier_mail_setting_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_supplier_mail_setting_name(select_supplier_mail_setting,edit_supplier_mail_setting_name)
                    supplier_mail_setting_dict = my.get_supplier_mail_setting_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Supplier Mail Setting Content"):
            st.rerun()
        select_supplier_mail_setting = st.selectbox("Select Function", list(supplier_mail_setting_dict.keys()), key="Edit Supplier Mail Setting Content Selectbox")
        if select_supplier_mail_setting != None:
            data_list = supplier_mail_setting_dict[select_supplier_mail_setting_name]

            if select_supplier_mail_setting_name != None:
                subject = st.text_input("Subject", data_list[0], key="Edit Supplier Mail Setting Subject")
                content = st.text_area("Content", data_list[1], key="Edit Supplier Mail Setting Content")

            if st.button('Confirm', key="Edit Supplier Mail Setting Content Button"):
                my.edit_supplier_mail_setting_row(select_supplier_mail_setting,subject,content)
                supplier_mail_setting_dict = my.get_supplier_mail_setting_dict()
                st.success("Change Successful")

def setting_check():
    st.title("Check")
    st.markdown('---')

    tab1, tab2, tab3, tab4 = st.tabs(["Show", "Create", "Edit Name", "Edit Content"])

    check_dict = my.get_check_dict()

    with tab1:
        if st.button('Refresh', key="Refresh Show Check"):
            st.rerun()
        select_check_name = st.selectbox("Select Check", list(check_dict.keys()), key="Show Check Selectbox")

        if check_dict != {} and select_check_name != None:
            data_str = check_dict[select_check_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_check_name != None:
                for i in range(0, 10):
                    a1, a2, a3 = st.columns([1, 2, 1])
                    with a1:
                        st.text_input("Check Column",data_df.loc[i,"Check Column"], disabled=True, key="Show Check Check Column {}".format(i))
                    with a2:
                        st.text_input("Formula",data_df.loc[i,"Formula"], disabled=True, key="Show Check Formula {}".format(i))
                    with a3:
                        st.text_input("Equal Column",data_df.loc[i,"Equal Column"], disabled=True, key="Show Check Equal Column {}".format(i))

    with tab2:
        check_column_list = []
        formula_list = []
        equal_column_list = []

        new_check_name = st.text_input("Name", key="Create Check Name")

        for i in range(0,10):
            a1, a2, a3 = st.columns([1, 2, 1])
            with a1:
                check_column_list.append(st.text_input("Check Column",key="Create Check Check Column {}".format(i)))
            with a2:
                formula_list.append(st.text_input("Formula",key="Create Check Formula {}".format(i)))
            with a3:
                equal_column_list.append(st.text_input("Equal Column",key="Create Check Equal Column {}".format(i)))

        if st.button('Confirm', key="Add Check"):
            if new_check_name in list(check_dict.keys()):
                st.warning("Duplicate Name")
            else:
                my.create_check(new_check_name,check_column_list,formula_list,equal_column_list)
                check_dict = my.get_check_dict()
                st.success("Added Successfully")

    with tab3:
        if st.button('Refresh', key="Refresh Edit Check Name"):
            st.rerun()
        select_check_name = st.selectbox("Select Check", list(check_dict.keys()), key="Edit Check Name Selectbox")

        edit_check_name = st.text_input("Edit Check Name")

        if select_check_name != None:
            if st.button('Confirm', key="Edit Requirements Introduction Name"):
                if edit_check_name in list(check_dict.keys()):
                    st.warning("Duplicate Name")
                else:
                    my.edit_check_name(select_check_name,edit_check_name)
                    check_dict = my.get_check_dict()
                    st.success("Change Successful")

    with tab4:
        if st.button('Refresh', key="Refresh Edit Check Content"):
            st.rerun()
        select_check_name = st.selectbox("Select Check", list(check_dict.keys()), key="Edit Check Content Selectbox")
        if check_dict != {} and select_check_name != None:
            data_str = check_dict[select_check_name]
            data_df = pd.DataFrame(json.loads(data_str)["data"], columns=json.loads(data_str)["columns"], index=json.loads(data_str)["index"])

            if select_check_name != None:
                for i in range(0, 10):
                    a1, a2, a3 = st.columns([1, 2, 1])
                    with a1:
                        st.text_input("Check Column", data_df.loc[i, "Check Column"], key="Edit Check Check Column {}".format(i))
                    with a2:
                        st.text_input("Formula", data_df.loc[i, "Formula"], key="Edit Check Formula {}".format(i))
                    with a3:
                        st.text_input("Equal Column", data_df.loc[i, "Equal Column"], key="Edit Check Equal Column {}".format(i))

                if st.button('Confirm', key="Edit Check Content"):
                    check_data = data_df.to_json(orient="split")
                    my.edit_check_row(select_check_name, check_data)
                    check_dict = my.get_check_dict()
                    st.success("Change Successful")

def download_form():
    st.title("Download")

    project_name_to_data, project_id_to_data = my.get_project_dict()
    project_name_list = list(project_name_to_data.keys())
    select_project = st.selectbox(label="Project", options=project_name_list)
    if select_project != None:
        project_data_list = project_name_to_data[select_project]
        project_id = project_data_list[0]
        title_id = project_data_list[2]
        requirements_introduction_id = project_data_list[3]
        common_columns_id = project_data_list[4]
        individual_columns_id = project_data_list[5]
        function_id = project_data_list[6]
        verification_code_id = project_data_list[7]
        internal_mail_receipients_id = project_data_list[8]
        supplier_mail_setting_id = project_data_list[9]
        setting_check_id = project_data_list[10]



        st.markdown("<font class='title-font'>Selected Program : {}</font>".format(select_project), unsafe_allow_html=True)
        st.write("")

        c1, c2, c3, c4 = st.columns([1, 1, 1, 1])
        with c1:
            st.write("Row Data")

            if st.button('Execute', key="Execute Row Data"):
                # 讀取Requirements Introduction
                requirements_introduction_data = my.get_requirements_introduction_by_id(requirements_introduction_id)
                product_list = [requirements_introduction_data[0], requirements_introduction_data[2], requirements_introduction_data[4],
                                requirements_introduction_data[6], requirements_introduction_data[8]]
                product_list = [i for i in product_list if i != ""]
                product_num = len(product_list)
                introduction_list = [requirements_introduction_data[requirements_introduction_data.index(i) + 1] for i in product_list]
                introduction_len = 1

                # 讀取Title
                title_list = my.get_title_by_id(title_id)
                title_list = [i for i in title_list if i != ""]
                title_len = len(title_list)

                # 讀取Check
                check_str = my.get_check_by_id(setting_check_id)
                check_df = pd.DataFrame(json.loads(check_str)["data"], columns=json.loads(check_str)["columns"], index=json.loads(check_str)["index"])
                check_df = check_df[check_df["Check Column"] != ""]
                check_df.reset_index(inplace=True,drop=True)

                # 讀取Results
                results_df_dict = get_product_results_df(project_id, product_list)

                wb = op.Workbook()
                for ind in range(0,len(product_list)):
                    product = product_list[ind]
                    sheetname = product.replace("/", " ").replace("\\", " ").replace("?", " ").replace("*", " ").replace("[", " ").replace("]", " ")  # Excel Sheetname 不接受
                    ws = wb.create_sheet(title=sheetname, index=ind)

                    results_df, different_list = add_formula_col(results_df_dict[product], check_df)
                    results_df.set_index("Product",drop=True,inplace=True)
                    try:
                        results_df.drop(columns=["index"],inplace=True)
                    except:
                        pass

                    for i in range(0,len(title_list)):
                        ws.cell(i+1, 1).value = title_list[i]

                    ws.cell(1 + title_len, 1).value = introduction_list[ind]

                    title_introduction_len = title_len + introduction_len + 1  # 多格一行

                    # 複製標題
                    for col in range(0, len(results_df.columns)):
                        ws.cell(1 + title_introduction_len, col + 1).value = results_df.columns[col]
                    # 複製資料
                    for ind in range(0, len(results_df)):
                        for col in range(0, len(results_df.columns)):
                            ws.cell(ind + 1 + 1 + title_introduction_len, col + 1).value = results_df.iloc[ind, col]

                    # 調整寬度
                    for i in range(0, len(list(ws.columns))):
                        col = list(ws.columns)[i]
                        column = col[0].column_letter  # Get the column name
                        if i == 0:
                            ws.column_dimensions[column].width = 80  # 先設定固定長度
                        else:
                            ws.column_dimensions[column].width = 15

                towrite = BytesIO()
                wb.save(towrite)
                towrite.seek(0)  # reset pointer
                st.download_button(label="Download", data=towrite.getvalue(), file_name="Row Data.xlsx", mime="application/vnd.ms-excel")

        with c2:
            st.write("Compared Form")
            if st.button('Execute', key="123"):
                # 讀取Requirements Introduction
                requirements_introduction_data = my.get_requirements_introduction_by_id(requirements_introduction_id)
                product_list = [requirements_introduction_data[0], requirements_introduction_data[2], requirements_introduction_data[4],
                                requirements_introduction_data[6], requirements_introduction_data[8]]
                product_list = [i for i in product_list if i != ""]
                product_num = len(product_list)
                introduction_list = [requirements_introduction_data[requirements_introduction_data.index(i) + 1] for i in product_list]
                introduction_len = 1

                # 讀取Title
                title_list = my.get_title_by_id(title_id)
                title_list = [i for i in title_list if i != ""]
                title_len = len(title_list)

                # 讀取Check
                check_str = my.get_check_by_id(setting_check_id)
                check_df = pd.DataFrame(json.loads(check_str)["data"], columns=json.loads(check_str)["columns"], index=json.loads(check_str)["index"])
                check_df = check_df[check_df["Check Column"] != ""]
                check_df.reset_index(inplace=True, drop=True)

                # 讀取Results
                results_df_dict = get_product_results_df(project_id, product_list)

                sty1 = NamedStyle(name="sty1", font=Font(name='Calibra', size=11, color='000000', bold=True), alignment=Alignment(wrap_text=True, horizontal='left', vertical='center'))
                sty2 = NamedStyle(name="sty2", font=Font(name='Calibra', size=11, color='000000'), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))
                sty3 = NamedStyle(name="sty3", font=Font(name='Calibra', size=11, color='000000'), alignment=Alignment(wrap_text=True, horizontal='left', vertical='center'))
                sty4 = NamedStyle(name="sty4", font=Font(name='Calibra', size=11, color='FF0000'), alignment=Alignment(wrap_text=True, horizontal='center', vertical='center'))

                wb = op.Workbook()
                for ind in range(0,len(product_list)):
                    product = product_list[ind]
                    sheetname = product.replace("/", " ").replace("\\", " ").replace("?", " ").replace("*", " ").replace("[", " ").replace("]", " ")  # Excel Sheetname 不接受
                    ws = wb.create_sheet(title=sheetname, index=ind)

                    results_df, different_list = add_formula_col(results_df_dict[product], check_df)
                    results_df.set_index("Product", drop=True, inplace=True)
                    try:
                        results_df.drop(columns=["index"], inplace=True)
                    except:
                        pass

                    campared_df = results_df.T

                    for i in range(0, title_len):
                        ws.cell(i + 1, 1).value = title_list[i]
                        ws.cell(i + 1, 1).style = sty3

                    ws.cell(1 + title_len, 1).value = introduction_list[ind]
                    ws.cell(1 + title_len, 1).style = sty3

                    title_introduction_len = title_len + introduction_len + 1  # 多格一行

                    # 複製標題
                    for ind in range(0, len(results_df.columns)):
                        ws.cell(ind + 1 + title_introduction_len, 1).value = results_df.columns[ind]
                        ws.cell(ind + 1 + title_introduction_len, 1).style = sty1

                    datetime_type = type(datetime.today().date())

                    # 複製資料
                    for ind in range(0, len(campared_df)):
                        col_name = campared_df.index[ind]

                        for col in range(0, len(campared_df.columns)):
                            row_id = campared_df.columns[col]
                            values = campared_df.iloc[ind, col]

                            if col_name == "Update DateTime":
                                values = str(values)
                            if type(values) == datetime_type:
                                values = str(values)
                            ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).value = values

                            if col_name in [i[1] for i in different_list if row_id == i[0]]:
                                # Check Data 不符合
                                ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).style = sty4
                            else:
                                ws.cell(ind + 1 + title_introduction_len, col + 1 + 1).style = sty2

                    # 調整寬度
                    for i in range(0, len(list(ws.columns))):
                        col = list(ws.columns)[i]
                        column = col[0].column_letter  # Get the column name
                        if i == 0:
                            ws.column_dimensions[column].width = 80  # 先設定固定長度
                        else:
                            ws.column_dimensions[column].width = 20

                towrite = BytesIO()
                wb.save(towrite)
                towrite.seek(0)  # reset pointer
                st.download_button(label="Download", data=towrite.getvalue(), file_name="Compared.xlsx", mime="application/vnd.ms-excel")

        with c3:
            st.write("Download Attachment")

def login():

    if st.session_state["authentication_status"]:
        authenticator.logout()
    elif st.session_state["authentication_status"] is False:
        authenticator.login()
        st.error('Username/password is incorrect')
    elif st.session_state["authentication_status"] is None:
        authenticator.login()


def get_product_results_df(project_id, product_list):
    results_df = my.get_results_by_project(project_id)
    results_df_list = [pd.DataFrame(json.loads(results_df.loc[i, "Data"])) for i in range(0, len(results_df))]

    product_df_dict = {}
    for product in product_list:
        product_df_dict[product] = pd.DataFrame()

    for results_df in results_df_list:
        for product in product_list:
            if len(product_df_dict[product]) == 0:
                product_df_dict[product] = results_df[results_df["Product"] == product]
            else:
                product_df_dict[product] = pd.concat([product_df_dict[product],results_df[results_df["Product"] == product]], ignore_index=True)

    return product_df_dict

def add_formula_col(results_df,check_df):
    if len(results_df) == 0:
        return results_df,[]
    else:
        results_df.reset_index(inplace=True, drop=False)

        different_list = []

        for i in range(0, len(results_df)):
            row_id = results_df.loc[i,"RowID"]
            for r in range(0, len(check_df)):
                check_column_name = check_df.loc[r, "Check Column"]
                equal_column = check_df.loc[r, "Equal Column"]
                formula = check_df.loc[r, "Formula"]
                for col in results_df.columns:
                    value = str(results_df.loc[i, col])
                    if value == "nan":
                        value = "0"
                    try:
                        # % 處理方式
                        if "%" in value:
                            value = str(float(value.replace("%","").strip()) / 100)
                    except:
                        pass
                    formula = formula.replace("[[" + col + "]]", value)
                try:
                    new_value = eval(formula)
                except Exception as error:
                    print(error)
                    new_value = 0
                results_df.loc[i, check_column_name] = new_value
                keyin_value = results_df.loc[i, equal_column]

                try: # 可能有問題，非數字
                    if math.isclose(float(keyin_value), new_value) == False:
                        different_list.append([row_id,check_column_name])
                except Exception as error:
                    print(error)
                    different_list.append([row_id, check_column_name])

        results_df.set_index('RowID', inplace=True)
        return results_df,different_list

page_names_to_funcs = {"Login":login,"Create Project": create_project,"Setting Title": setting_title,
                       "Setting Requirements Introduction": setting_requirements_introduction,
                       "Setting Common Columns":setting_common_columns,
                       "Setting Individual Columns":setting_individual_columns,
                       "Setting Function":setting_function,
                       "Setting Verification Code":setting_verification_code,
                       "Setting Internal Mail Receipients":setting_internal_mail_receipients,
                       "Setting Supplier Mail":setting_supplier_mail,
                       "Setting Check":setting_check,
                       "Download":download_form}

functions_name = st.sidebar.selectbox("Option", page_names_to_funcs.keys())

page_names_to_funcs[functions_name]()
